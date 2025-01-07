"""
Title: Read a Snellius Excel report file and enrich with AD info
Author: Peter Vos
version: 0.8-alpha

Usage (Local): python3 get_user_data.py

Requires: openpyxl (pip install openpyxl), ldap3 (pip install ldap3)

(C) Peter J.M. Vos, VU Amsterdam, 2024. Licenced for use under the BSD 3 clause

Make sure to set the AD credentials in ./config.py. Copy the Excel file you want to process in ./data and set the filename at the bottom
"""

from ldap3 import Server, Connection, ALL, NTLM, ALL_ATTRIBUTES
from config import AD_PW, AD_USER, AD_SERVER, REPORT_PATH
import openpyxl
import json

def ad_lookup(email):
    # we could try to add a lookup by first/last name for the other addresses later
    print(email)
    conn.search('dc=vu,dc=local', f'(&(objectclass=person)(|(proxyaddresses=SMTP:{email})(proxyaddresses=smtp:{email})))',
        attributes=['department','company','eduPersonAffiliation','title','displayName'])
    data={}
    try: 
        entry = conn.entries[0]
        data = {
            'department': "|".join(entry.department),
            'company': "|".join(entry.company),
            'eduPersonAffiliation': "|".join(entry.eduPersonAffiliation),
            'title':  "|".join(entry.title),
            'displayName':  "|".join(entry.displayName),
            'account': account
        }
    except IndexError: # not found in AD
        pass
    return data

def get_data(reportfile):
    reportdate = reportfile.split('.')[1]
    workbook = openpyxl.load_workbook(f"{reportfile}")
    sheet = workbook.active

    data = {}
    filename = f'snellius_usersAD-{reportdate}.json'
    datafile = f'{REPORT_PATH}/{filename}'
    data = {}

    server = Server(AD_SERVER, use_ssl=True, get_info=ALL)

    with Connection(server, AD_USER, AD_PW, auto_bind=True) as conn:
        for i in range(1, sheet.max_row+1):
            print(i)
            code = sheet.cell(row=i, column=1).value
            description = sheet.cell(row=i, column=2).value
            #  use the "sub-heading"-rows in the decription column to set the type of budget 
            if description == 'Snellius VU CPU-compute 2024':
                budget_type = 'CPU'
            elif  description == 'Snellius VU GPU-compute 2024':
                budget_type = 'GPU'
            elif description == 'Snellius VU project storage 2024':
                budget_type = 'projectspace'
            # TODO read the column headers, they tend to change
            account = sheet.cell(row=i, column=5).value
            email = sheet.cell(row=i, column=9).value
            budget = sheet.cell(row=i, column=13).value
            usage = sheet.cell(row=i, column=14).value
            if '@' in email and code.startswith('2307090'): # only interested in the "totals" rows with an email address 
                if email not in data:
                    data[email] = {
                        'account': account
                    }
                    if email.endswith(("vu.nl", "acta.nl")): # VU users
                        data[email]=ad_lookup(email)
                if budget_type in data[email]:
                        data[email][budget_type]['budget'] += budget
                        data[email][budget_type]['usage'] += usage
                else:
                    data[email][budget_type] = {
                        'budget': budget,
                        'usage': usage
                    }

    # dump the data in a json for now
    with open(datafile, 'w') as fp:
        json.dump(data, fp)



if __name__ == '__main__':
    get_data(reportfile = f"{REPORT_PATH}/2307090_24.20250106.xlsx")