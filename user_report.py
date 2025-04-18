"""
Title: Converts a Snellius excel report to json, looks up email addresses in the AD and outputs to a per-user xlsx file
Author: Peter Vos
version: 1.0

Usage (Local): python3 user_report.py

Requires: openpyxl (pip install openpyxl)

(C) Peter J.M. Vos, VU Amsterdam, 2025. Licenced for use under the BSD 3 clause
"""

import openpyxl
from datetime import datetime
import json
from config import REPORT_PATH
from create_json import create_json_from_report
from ad_lookup import ad_lookup


def create_excel(datafile, years):
    """
    Converts a json-converted Snellius report to excel, outputs a per-user xlsx file.

    Parameters
    ----------
    datafile : str
        Path to the input json file.
    years : list of int
        List of years for which the reports should be generated.

    Returns
    -------
    None

    Notes
    -----
    This function assumes that the input json file was created by the create_json.py and the ad_lookup.py scripts.
    
    The output will be a per-user xlsx file.
    
    Example
    -------
    >>> create_excel('input.xlsx', [2022, 2023])
    """


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "snellius_usage"

    with open(f"{datafile}", "r+") as fp:
        data = json.load(fp)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "email"
    sheet.cell(row=1, column=2).value = "account"
    sheet.cell(row=1, column=3).value = "company"
    sheet.cell(row=1, column=4).value = "department"
    sheet.cell(row=1, column=5).value = "eduPersonAffiliation"
    sheet.cell(row=1, column=6).value = "title"
    sheet.cell(row=1, column=7).value = "displayName"
    sheet.cell(row=1, column=8).value = "cpu budget"
    sheet.cell(row=1, column=9).value = "cpu usage total"
    sheet.cell(row=1, column=10).value = "gpu budget"
    sheet.cell(row=1, column=11).value = "gpu usage total"
    sheet.cell(row=1, column=12).value = "projectspace usage"

    c = 13
    for year in years:
        sheet.cell(row=1, column=c).value = f"{year} cpu"
        sheet.cell(row=1, column=c + 1).value = f"{year} gpu"
        sheet.cell(row=1, column=c + 2).value = f"{year} projectspace"
        c += 3
    r = 2
    for account, userdata in data.items():
        projectspace = userdata.get("projectspace", {})
        CPU = userdata.get("CPU", {})
        GPU = userdata.get("GPU", {})
        projecspace = userdata.get("projectspace", {})
        addata = userdata.get("AD", {})

        sheet.cell(row=r, column=1).value = userdata.get("email", "")
        sheet.cell(row=r, column=2).value = account
        sheet.cell(row=r, column=3).value = addata.get("company", "")
        sheet.cell(row=r, column=4).value = addata.get("department", "")
        sheet.cell(row=r, column=5).value = addata.get("eduPersonAffiliation", "")
        sheet.cell(row=r, column=6).value = addata.get("title", "")
        sheet.cell(row=r, column=7).value = addata.get("displayName", "")
        sheet.cell(row=r, column=8).value = CPU.get("budget", 0)
        sheet.cell(row=r, column=9).value = round(CPU.get("total_usage", 0))
        sheet.cell(row=r, column=10).value = GPU.get("budget", 0)
        sheet.cell(row=r, column=11).value = round(GPU.get("total_usage", 0))
        sheet.cell(row=r, column=12).value = projectspace.get("total_usage", 0)
        c = 13
        for year in years:
            sheet.cell(row=r, column=c).value = round(CPU.get(year, 0), 0)
            sheet.cell(row=r, column=c + 1).value = round(GPU.get(year, 0), 0)
            sheet.cell(row=r, column=c + 2).value = round(projectspace.get(year, 0), 0)
            c += 3
        r += 1
    filename = datafile.replace(".json", ".xlsx")
    workbook.save(filename)
    return filename


if __name__ == "__main__":
    defreport = "2307090_23.20240705.xlsx"
    reportfile = input(f"Which xlsx report do you want to process [{defreport}]? (file must be in path {REPORT_PATH})") or defreport
    defcols = "1"
    ignorecols = input(f"How many of the rightmost columns do you want to ignore [{defcols}]? (eg choose 1 if the last columns contain info on just the first day of the current month)") or defcols
    print(f"Start processing {REPORT_PATH}/{reportfile}, ignore last {ignorecols} columns.")
    datafile, years = create_json_from_report(
        reportfile=reportfile, ignorecol=int(ignorecols)
    )
    print(f'Created {datafile}')
    ad_datafile = ad_lookup(datafile, lookup=True)
    print(f'Added AD info in {ad_datafile}')
    filename = create_excel(ad_datafile, years)
    print(f'Output written to {filename}')
