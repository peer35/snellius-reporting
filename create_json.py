from config import REPORT_PATH
import openpyxl
import json


def get_headings(sheet):
    """
    Extracts headings from a given sheet in an Excel workbook.
    
    Args:
        sheet (openpyxlWorksheet): The sheet to extract headings from.
        
    Returns:
        list: A list of heading values.
    """
    headings = []
    for n in range(1, sheet.max_column + 1):
        headings.append(sheet.cell(1, column=n).value)
    return headings


def create_json_from_report(reportfile, ad_lookup=False, ignorecol=1):
    """
    Creates a JSON file from the given report Excel file.
    
    Args:
        reportfile (str): The path to the Excel report file.
        ad_lookup (bool): Whether to perform an AD lookup. Defaults to False.
        ignorecol (int): The column number to ignore when processing data. Defaults to 1.
        
    Returns:
        dict: Filename of json file containing the processed data.
    """
    # Code remains unchanged

    reportdate = reportfile.split(".")[1]
    workbook = openpyxl.load_workbook(f"{REPORT_PATH}/{reportfile}")
    sheet = workbook.active
    headings = get_headings(sheet)
    col_account = headings.index("Account") + 1
    try:
        col_email = headings.index("email") + 1
    except ValueError:
        col_email = False
    col_usage = headings.index("SrvUsage") + 1
    col_budget = headings.index("Budget") + 1
    col_trend = (
        headings.index("trend") + 1
    )  # should be the last column before the monthly usage columns

    data = {}
    years = []  # years found in excel report
    for i in range(1, sheet.max_row + 1):
        code = sheet.cell(row=i, column=1).value
        description = sheet.cell(row=i, column=2).value
        #  use the "sub-heading"-rows in the decription column to set the type of budget
        if description.startswith("Snellius VU CPU-compute "):
            budget_type = "CPU"
        elif description.startswith("Snellius VU GPU-compute "):
            budget_type = "GPU"
        elif description.startswith("Snellius VU project storage "):
            budget_type = "projectspace"
        account = sheet.cell(row=i, column=col_account).value
        #if account.startswith("snel-vusr") and code.startswith("2307090"):
        if account.startswith("snel-") and code.startswith("2307090"):
            if account not in data:
                data[account] = {}
            budget = sheet.cell(row=i, column=col_budget).value
            usage = sheet.cell(row=i, column=col_usage).value
            if col_email:
                email = sheet.cell(row=i, column=col_email).value
                if email not in data:
                    data[account]["email"] = email
            if budget_type in data[account]:
                data[account][budget_type]["budget"] += budget
                data[account][budget_type]["total_usage"] += usage
            else:
                data[account][budget_type] = {"budget": budget, "total_usage": usage}
            # add usage per year using the per month totals
            for j in range(col_trend + 1, sheet.max_column + 1 - ignorecol):
                # IGNORES THE LAST MONTH! This contains only info on the first few days.
                # In january 2026 we can add the 2025 totals of the July 2025 report to the 2025 totals of the January 2026 report for a full year report.
                # Project space: months 5TB, 3 months 10TB = 40TBmonth?? Then if 1TB for a year = €480.0 costs are 40*(480/12)=1600
                datestr = headings[j - 1]
                print(f'***** {datestr}')
                year = datestr[0:4]
                if year not in years:
                    years.append(year)
                if sheet.cell(row=i, column=j).value == "":
                    month_usage = 0
                else:
                    month_usage = sheet.cell(row=i, column=j).value
                print(account,datestr,year,month_usage, budget_type)
                if year in data[account][budget_type]:
                    data[account][budget_type][year] += month_usage
                else:
                    data[account][budget_type][year] = month_usage   
                data[account][budget_type][datestr] = month_usage

    # dump the data in a json for now
    #datafile = f"{REPORT_PATH}/processed/{reportfile.replace('.xlsx','.json')}"
    datafile = f"data/{reportfile.replace('.xlsx','.json')}"
    print(datafile)
    with open(datafile, "w") as fp:
        json.dump(data, fp)
    return datafile, years


if __name__ == "__main__":
    datafile, years = create_json_from_report(reportfile="2307090_25.20251006.xlsx")
