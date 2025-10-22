import openpyxl
import json

# Months to calculate over
datestrings = ["2024-11","2024-12","2025-01","2025-02","2025-03","2025-04","2025-05","2025-06","2025-07","2025-08","2025-09","2025-10"]  
# Input data: e.g. july report over previous period + november report
#datafiles = ["data/2307090_24.20250707_AD.json","data/2307090_25.20251006_AD.json"]
datafiles = ["data/2307090_25.20251006_AD.json"]

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = "email"
sheet.cell(row=1, column=2).value = "account"
sheet.cell(row=1, column=3).value = "company"
sheet.cell(row=1, column=4).value = "department"
sheet.cell(row=1, column=5).value = "cpu usage total"
sheet.cell(row=1, column=6).value = "gpu usage total"
sheet.cell(row=1, column=7).value = "projectspace usage"

with open("data/userdata.json", "r") as fu:
    addata=json.load(fu)

accounts=0
row = 1
output = {}
for datafile in datafiles:

    with open(f"{datafile}", "r+") as fp:
        data = json.load(fp)
        for account, userdata in data.items():
            email = userdata["email"]
            if account not in output:
                if (
                    email in addata
                ):  # use info found previously in case account has been deleted
                    accountdata = addata[email]
                else:
                    accountdata = {}
                output[account] = {
                    "email": email,
                    "company": accountdata.get("company",""),
                    "department": accountdata.get("department",""),
                    "CPU_total": 0,
                    "GPU_total": 0,
                    "project_space": 0,
                }
                accounts += 1
            projectspace = userdata.get("projectspace", {})
            CPU = userdata.get("CPU", {})
            GPU = userdata.get("GPU", {})

            for entry in CPU:   
                if entry in datestrings:
                    output[account]["CPU_total"] = output[account]["CPU_total"] + CPU[entry] 
            for entry in GPU:
                if entry in datestrings:
                    output[account]["GPU_total"] = output[account]["GPU_total"] + GPU[entry]
            ps = userdata.get("projectspace",{})
            output[account]["project_space"] = ps.get("budget", 0)

with open("data/snellius_usage2025.json", "w") as fp:
    json.dump(output, fp)

row = 2
for account, reportdata in output.items():
    sheet.cell(row=row, column=1).value = reportdata["email"]
    sheet.cell(row=row, column=2).value = account
    sheet.cell(row=row, column=3).value = reportdata["company"]
    sheet.cell(row=row, column=4).value = reportdata["department"]
    sheet.cell(row=row, column=5).value = reportdata["CPU_total"]
    sheet.cell(row=row, column=6).value = reportdata["GPU_total"]
    sheet.cell(row=row, column=7).value = reportdata["project_space"]
    row += 1

print(accounts)
workbook.save("data/snellius_usage2025.xlsx")